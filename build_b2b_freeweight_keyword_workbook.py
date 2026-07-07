from pathlib import Path
import csv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


ROOT = Path(r"C:\Users\Kloe\Documents\chinafreeweight独立站")
OUT = ROOT / "outputs" / "b2b_freeweight_keywords"
OUT.mkdir(parents=True, exist_ok=True)

HEADERS = [
    "keyword",
    "keyword_type",
    "category",
    "intent",
    "buyer_stage",
    "search_volume_estimate",
    "difficulty_estimate",
    "commercial_value_score",
    "seo_score",
    "country_target",
    "recommended_page_type",
    "recommended_url",
    "recommended_title",
    "recommended_h1",
    "recommended_meta_description",
    "priority_score",
    "notes",
]

SEEDS = [
    {"key": "dumbbell", "slug": "dumbbells", "cat": "Dumbbells"},
    {"key": "weight plates", "slug": "weight-plates", "cat": "Weight Plates"},
    {"key": "bumper plates", "slug": "bumper-plates", "cat": "Bumper Plates"},
    {"key": "rubber dumbbells", "slug": "rubber-dumbbells", "cat": "Rubber Dumbbells"},
]

PRODUCTS = {
    "dumbbell": [
        "commercial dumbbells", "fixed dumbbells", "hex dumbbells", "rubber hex dumbbells",
        "urethane dumbbells", "round head dumbbells", "studio dumbbells", "dumbbell set with rack",
        "custom logo dumbbells", "kg dumbbell set", "lb dumbbell set",
    ],
    "weight plates": [
        "olympic weight plates", "rubber coated weight plates", "urethane weight plates",
        "cast iron weight plates", "grip weight plates", "tri grip weight plates",
        "calibrated weight plates", "steel weight plates", "custom logo weight plates",
        "commercial weight plates",
    ],
    "bumper plates": [
        "olympic bumper plates", "competition bumper plates", "training bumper plates",
        "crumb rubber bumper plates", "virgin rubber bumper plates", "colored bumper plates",
        "black bumper plates", "kg bumper plates", "lb bumper plates", "custom logo bumper plates",
    ],
    "rubber dumbbells": [
        "commercial rubber dumbbells", "rubber hex dumbbells", "rubber coated dumbbells",
        "fixed rubber dumbbells", "round rubber dumbbells", "rubber dumbbell set",
        "rubber dumbbells with rack", "custom logo rubber dumbbells", "odorless rubber dumbbells",
        "premium rubber dumbbells",
    ],
}

PURCHASE = ["supplier", "manufacturer", "factory", "wholesale", "bulk", "for sale", "price", "cost", "quote", "OEM", "ODM", "private label", "exporter", "import from China"]
APPLICATIONS = ["commercial gyms", "gym chains", "fitness centers", "hotels and resorts", "schools and universities", "apartment gyms", "military gyms", "sports performance centers"]
COUNTRIES = ["United States", "United Kingdom", "Canada", "Brazil", "Mexico", "Australia", "Spain", "Germany", "France", "UAE", "Saudi Arabia", "South Africa", "Chile", "Colombia", "India"]
TECHNICAL = [
    ("rubber odor", "Material / quality"), ("shore hardness", "Material / quality"),
    ("stainless steel handle", "Specification"), ("knurled handle", "Specification"),
    ("drop test", "Durability"), ("weight tolerance", "Specification"),
    ("IWF standard", "Standard"), ("Olympic 2 inch hole", "Specification"),
    ("kg vs lb", "Specification"), ("custom logo", "Branding"),
    ("packaging for export", "Logistics"), ("container loading quantity", "Logistics"),
]
QUESTIONS = [
    "how to choose {seed} for a commercial gym",
    "what is the difference between rubber and urethane {seed}",
    "what are the best {seed} for gym chains",
    "how much do commercial {seed} cost",
    "what specifications matter when buying {seed} in bulk",
    "how to compare {seed} manufacturers",
    "what is a good MOQ for {seed} wholesale",
    "how to inspect {seed} quality before shipment",
    "how to maintain {seed} in fitness centers",
    "what is the lifespan of commercial {seed}",
]
AI_QUERIES = [
    "Which China {seed} manufacturer is best for global B2B buyers?",
    "Can you compare {seed} suppliers for a new commercial gym project?",
    "What should a gym chain ask before ordering {seed} in bulk?",
    "Give me a buyer checklist for importing {seed} from China.",
    "What is the landed cost structure for wholesale {seed}?",
    "Which {seed} material is better for hotels, schools, and fitness centers?",
    "How can I evaluate {seed} quality without visiting the factory?",
    "What are the safest {seed} options for high traffic commercial gyms?",
    "Create an RFQ template for custom logo {seed}.",
    "What are the common problems with cheap {seed} and how to avoid them?",
]


def slugify(text):
    out = "".join(ch.lower() if ch.isalnum() else "-" for ch in text.replace("&", "and"))
    while "--" in out:
        out = out.replace("--", "-")
    return out.strip("-")


def title_case(text):
    return " ".join(w.upper() if w in {"OEM", "ODM", "RFQ", "MOQ", "IWF", "UAE"} else w.capitalize() for w in text.split())


def volume(keyword_type, keyword):
    if keyword_type == "Main Keywords":
        return "High" if len(keyword.split()) == 1 else "Medium-High"
    if keyword_type == "Product Keywords":
        return "Medium"
    if keyword_type == "Commercial Keywords":
        return "Medium" if any(x in keyword for x in ["price", "for sale"]) else "Low-Medium"
    if keyword_type in {"Question Keywords", "AI / GEO Keywords", "Country Keywords"}:
        return "Low-Medium"
    return "Low"


def difficulty(keyword_type, keyword):
    base = {
        "Main Keywords": 88 if len(keyword.split()) == 1 else 78,
        "Product Keywords": 62,
        "Commercial Keywords": 52,
        "Application Keywords": 38,
        "Technical Keywords": 32,
        "Question Keywords": 35,
        "Country Keywords": 42,
        "AI / GEO Keywords": 28,
    }.get(keyword_type, 45)
    if any(x in keyword.lower() for x in ["supplier", "manufacturer", "wholesale", "factory"]):
        base -= 5
    return max(18, min(92, base))


def commercial(keyword_type, keyword):
    base = {
        "Commercial Keywords": 92,
        "Application Keywords": 86,
        "Country Keywords": 84,
        "Product Keywords": 78,
        "Technical Keywords": 70,
        "Question Keywords": 62,
        "AI / GEO Keywords": 82,
        "Main Keywords": 72,
    }.get(keyword_type, 55)
    if any(x in keyword.lower() for x in ["price", "quote", "supplier", "manufacturer", "factory"]):
        base += 4
    return min(98, base)


def page_type(keyword_type, keyword):
    if keyword_type == "Main Keywords":
        return "Product Category Pages"
    if keyword_type == "Product Keywords":
        return "Product Detail Pages"
    if keyword_type == "Commercial Keywords":
        return "FAQ Pages" if any(x in keyword for x in ["price", "cost", "quote"]) else "Product Category Pages"
    if keyword_type == "Application Keywords":
        return "Application Pages"
    if keyword_type in {"Technical Keywords", "Question Keywords"}:
        return "Blog Articles"
    if keyword_type == "Country Keywords":
        return "Country Landing Pages"
    if keyword_type == "AI / GEO Keywords":
        return "FAQ Pages"
    return "Blog Articles"


def buyer_stage(keyword_type, keyword):
    if keyword_type == "Commercial Keywords" and any(x in keyword.lower() for x in ["quote", "price", "for sale", "supplier", "manufacturer", "factory"]):
        return "Decision"
    if keyword_type in {"Application Keywords", "Country Keywords", "Product Keywords", "Main Keywords"}:
        return "Consideration"
    return "Research"


def url_for(keyword_type, keyword, seed_slug, country):
    s = slugify(keyword)
    if keyword_type == "Main Keywords":
        return f"/products/{seed_slug}/"
    if keyword_type == "Product Keywords":
        return f"/products/{seed_slug}/{s}/"
    if keyword_type == "Commercial Keywords":
        return f"/faq/{s}/" if any(x in keyword for x in ["price", "cost", "quote"]) else f"/products/{seed_slug}/{s}/"
    if keyword_type == "Application Keywords":
        return f"/applications/{s}/"
    if keyword_type in {"Technical Keywords", "Question Keywords"}:
        return f"/blog/{s}/"
    if keyword_type == "Country Keywords":
        return f"/markets/{slugify(country)}/{seed_slug}/"
    if keyword_type == "AI / GEO Keywords":
        return f"/faq/{s}/"
    return f"/blog/{s}/"


def meta(keyword, keyword_type):
    if keyword_type == "Commercial Keywords":
        return f"Source {keyword} for commercial gyms, gym chains, hotels, schools, and distributors. Request OEM options, bulk pricing, lead time, and export packaging."
    if keyword_type == "Application Keywords":
        return f"Explore free weight solutions for {keyword.replace(' for ', ' in ')}. Compare dumbbells, plates, racks, branding, and bulk supply options."
    if keyword_type == "Country Keywords":
        return f"B2B {keyword} supply for importers, distributors, and commercial gym projects. OEM branding, export packaging, and bulk quotation support."
    if keyword_type in {"AI / GEO Keywords", "Question Keywords"}:
        return f"Practical B2B guide to {keyword}. Learn specifications, supplier checks, pricing factors, quality control, and buying recommendations."
    return f"Commercial-grade {keyword} for fitness centers, gym chains, hotels, schools, and distributors. OEM branding, bulk supply, and export-ready packaging."


def make_row(keyword, keyword_type, category, seed_slug="free-weights", country="Global", notes=""):
    diff = difficulty(keyword_type, keyword)
    comm = commercial(keyword_type, keyword)
    vol_bonus = 15 if "High" in volume(keyword_type, keyword) else 10 if "Medium" in volume(keyword_type, keyword) else 6
    seo = round(comm * 0.45 + (100 - diff) * 0.35 + vol_bonus)
    priority = round(comm * 0.45 + seo * 0.35 + (12 if keyword_type in {"Commercial Keywords", "Application Keywords"} else 6))
    return {
        "keyword": keyword,
        "keyword_type": keyword_type,
        "category": category,
        "intent": "Informational + B2B qualification" if keyword_type in {"Question Keywords", "AI / GEO Keywords", "Technical Keywords"} else "Commercial investigation",
        "buyer_stage": buyer_stage(keyword_type, keyword),
        "search_volume_estimate": volume(keyword_type, keyword),
        "difficulty_estimate": diff,
        "commercial_value_score": comm,
        "seo_score": seo,
        "country_target": country,
        "recommended_page_type": page_type(keyword_type, keyword),
        "recommended_url": url_for(keyword_type, keyword, seed_slug, country),
        "recommended_title": f"{title_case(keyword)} | {'B2B Buying Guide' if keyword_type in {'Question Keywords', 'AI / GEO Keywords'} else 'Commercial Gym Equipment Manufacturer'}",
        "recommended_h1": title_case(keyword),
        "recommended_meta_description": meta(keyword, keyword_type),
        "priority_score": priority,
        "notes": notes or "Simulated from Google search logic, B2B buyer intent, and commercial gym procurement patterns.",
    }


rows = [
    make_row("commercial free weights", "Main Keywords", "Free Weight", notes="Home page primary B2B topic; high value but competitive."),
    make_row("free weight equipment manufacturer", "Main Keywords", "Free Weight", notes="Strong homepage/manufacturer positioning keyword."),
    make_row("commercial gym free weights", "Main Keywords", "Free Weight", notes="Best bridge between product category and B2B application intent."),
]

for seed in SEEDS:
    key = seed["key"]
    rows.append(make_row(key, "Main Keywords", seed["cat"], seed["slug"], notes="Broad head term; useful for category relevance but high SEO difficulty."))
    rows.append(make_row(f"commercial {key}", "Main Keywords", seed["cat"], seed["slug"], notes="B2B-modified main term for category landing page."))
    for product in PRODUCTS[key]:
        rows.append(make_row(product, "Product Keywords", seed["cat"], seed["slug"]))
    for mod in PURCHASE:
        rows.append(make_row(f"{key} {mod}", "Commercial Keywords", seed["cat"], seed["slug"], notes="High inquiry value; use RFQ, MOQ, customization, and shipping proof."))
    for app in APPLICATIONS:
        rows.append(make_row(f"{key} for {app}", "Application Keywords", seed["cat"], seed["slug"], notes="Application page should show project use cases, recommended sets, and inquiry CTA."))
    for term, cat in TECHNICAL:
        rows.append(make_row(f"{key} {term}", "Technical Keywords", cat, seed["slug"], notes="Supporting content for buyer education and internal links to product pages."))
    for tpl in QUESTIONS:
        rows.append(make_row(tpl.format(seed=key), "Question Keywords", seed["cat"], seed["slug"], notes="Use concise answer blocks, FAQ schema style formatting, and product-page internal links."))
    for tpl in AI_QUERIES:
        rows.append(make_row(tpl.format(seed=key), "AI / GEO Keywords", seed["cat"], seed["slug"], notes="Write direct answers, comparison tables, buyer checklist, and RFQ CTA."))
    for country in COUNTRIES:
        rows.append(make_row(f"{key} supplier in {country}", "Country Keywords", seed["cat"], seed["slug"], country, notes="Country landing page should mention shipping, certifications, local buyer concerns, and RFQ."))
        rows.append(make_row(f"{key} manufacturer for {country} importers", "Country Keywords", seed["cat"], seed["slug"], country, notes="Import-focused B2B term for distributors and gym equipment buyers."))

dedup = {}
for row in rows:
    dedup.setdefault(row["keyword"].lower(), row)
rows = sorted(dedup.values(), key=lambda r: (-r["priority_score"], -r["commercial_value_score"], r["keyword"]))

competitor_patterns = [
    ["Common Title", "Product + modifier + commercial proof", "Rubber Hex Dumbbells | Commercial Gym Equipment Supplier", "Winning pages combine exact product term, material, and buyer segment."],
    ["Common Title", "Product + for sale / wholesale", "Bumper Plates for Sale | Wholesale Olympic Plates", "Retail-heavy SERPs use for-sale; B2B sites should add wholesale/manufacturer angle."],
    ["Common Meta Description", "Product range + quality + CTA", "Shop commercial-grade dumbbells, plates, and racks with bulk pricing, custom logo options, and export packaging.", "Best meta descriptions answer what, who for, and why contact now."],
    ["Common Meta Description", "Specification + use case", "Durable rubber coated free weights for gyms, fitness centers, hotels, and schools. Request a quote for bulk orders.", "Application terms distinguish B2B from retail pages."],
    ["Common H1", "Exact product category", "Commercial Rubber Dumbbells", "Most ranking pages use simple exact-match H1 without stuffing."],
    ["Common H2", "Product types", "Hex Dumbbells / Urethane Dumbbells / Dumbbell Sets / Dumbbell Racks", "Category pages should expose subcategories through H2 blocks."],
    ["Common H2", "Buying information", "Specifications / Custom Logo / Packaging / MOQ / Shipping / FAQ", "B2B pages need procurement content, not just product cards."],
    ["Image ALT", "Product + material + weight", "rubber hex dumbbell 20kg commercial gym", "Alt text usually names product, material, weight, and use context."],
    ["URL Structure", "Short product-category hierarchy", "/products/dumbbells/rubber-hex-dumbbells/", "Clean URLs usually perform better than parameter URLs."],
    ["URL Structure", "Application landing pages", "/applications/commercial-gyms/free-weights/", "Good for B2B buyer segments and internal linking."],
    ["Product Classification", "By product line", "Dumbbells / Weight Plates / Bumper Plates / Barbells / Racks", "Start navigation with buyer-recognized product categories."],
    ["Product Classification", "By material and standard", "Rubber / Urethane / Cast Iron / Olympic / KG / LB", "Filters or subcategory pages can support long-tail SEO."],
    ["Blog Structure", "Buying guide cluster", "How to choose dumbbells for a commercial gym", "High-value blog topics qualify buyers and link to category pages."],
    ["Blog Structure", "Comparison cluster", "Rubber vs urethane dumbbells; bumper plates vs weight plates", "Comparison posts are good for AIO/GEO visibility."],
    ["Blog Structure", "Procurement cluster", "MOQ, shipping, quality inspection, factory audit, RFQ checklist", "Important for importers and distributors near inquiry stage."],
    ["Blog Structure", "Maintenance cluster", "How to clean rubber dumbbells; how long bumper plates last", "Supports specification confidence and post-purchase topics."],
]

scoring_logic = [
    ["Search volume potential", "Estimated from head-term breadth and modifier depth, not real SEO tool data.", "High / Medium-High / Medium / Low-Medium / Low"],
    ["Commercial value score", "0-100 estimate of likely inquiry value and order size.", "Supplier, manufacturer, wholesale, quote, country, and application terms score highest."],
    ["SEO difficulty estimate", "0-100 estimate from SERP competitiveness logic.", "Broad head terms are hardest; long-tail B2B and technical terms are easier."],
    ["SEO score", "Composite of commercial value, lower difficulty, and volume potential.", "Higher means better organic opportunity."],
    ["Priority score", "Final action priority for B2B independent-site content planning.", "Balances inquiry value, SEO opportunity, and page suitability."],
    ["Independent page fit", "Mapped through recommended_page_type.", "Category/detail/application/country pages for conversion; blog/FAQ for education and AIO."],
]

site_map = [
    ["site_section", "target_keyword_types", "primary_goal", "recommended_url_pattern"],
    ["Home", "Main Keywords", "Position as commercial free weight manufacturer and export supplier", "/"],
    ["Product Category Pages", "Main Keywords, Commercial Keywords", "Rank for product families and supplier/manufacturer terms", "/products/{category}/"],
    ["Product Detail Pages", "Product Keywords", "Capture material, type, size, and custom logo long-tail terms", "/products/{category}/{product}/"],
    ["Application Pages", "Application Keywords", "Convert gyms, gym chains, hotels, schools, and fitness centers", "/applications/{segment}/"],
    ["Blog Articles", "Question Keywords, Technical Keywords", "Educate buyers and support AIO/GEO answer visibility", "/blog/{topic}/"],
    ["FAQ Pages", "AI / GEO Keywords, pricing and procurement questions", "Answer procurement questions and push RFQ action", "/faq/{question}/"],
    ["Country Landing Pages", "Country Keywords", "Target importers and distributors by country", "/markets/{country}/{category}/"],
]


def style_sheet(ws, table_name, widths):
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    max_col = ws.max_column
    max_row = ws.max_row
    ref = f"A1:{get_column_letter(max_col)}{max_row}"
    tab = Table(displayName=table_name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(tab)
    header_fill = PatternFill("solid", fgColor="1F4E5F")
    white = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="E5E7EB")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = white
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = Border(bottom=thin)
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.auto_filter.ref = ref


wb = Workbook()
ws = wb.active
ws.title = "Keyword Master"
ws.append(HEADERS)
for row in rows:
    ws.append([row[h] for h in HEADERS])
style_sheet(ws, "KeywordMaster", [42, 24, 22, 28, 16, 20, 16, 18, 12, 18, 24, 42, 44, 34, 58, 14, 58])
for row in ws.iter_rows(min_row=2, min_col=7, max_col=9):
    for cell in row:
        cell.number_format = "0"
for cell in ws["P"][1:]:
    cell.number_format = "0"

ws2 = wb.create_sheet("Competitor SEO Structure")
ws2.append(["seo_area", "pattern", "example", "b2b_takeaway"])
for row in competitor_patterns:
    ws2.append(row)
style_sheet(ws2, "CompetitorPatterns", [26, 38, 62, 72])

ws3 = wb.create_sheet("Scoring Logic")
ws3.append(["dimension", "logic", "scale"])
for row in scoring_logic:
    ws3.append(row)
style_sheet(ws3, "ScoringLogic", [30, 76, 66])

ws4 = wb.create_sheet("Site Structure Map")
for row in site_map:
    ws4.append(row)
style_sheet(ws4, "SiteStructureMap", [30, 44, 76, 40])

xlsx_path = OUT / "B2B_FreeWeight_SEO_Keyword_System.xlsx"
csv_path = OUT / "b2b_freeweight_keyword_master.csv"
wb.save(xlsx_path)

with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
    writer = csv.DictWriter(f, fieldnames=HEADERS)
    writer.writeheader()
    writer.writerows(rows)

check = load_workbook(xlsx_path, read_only=True, data_only=True)
assert "Keyword Master" in check.sheetnames
assert check["Keyword Master"].max_row == len(rows) + 1
assert check["Keyword Master"].max_column == len(HEADERS)

desktop = Path.home() / "Desktop"
if desktop.exists():
    desktop_xlsx = desktop / xlsx_path.name
    desktop_csv = desktop / csv_path.name
    desktop_xlsx.write_bytes(xlsx_path.read_bytes())
    desktop_csv.write_bytes(csv_path.read_bytes())

print(f"xlsx={xlsx_path}")
print(f"csv={csv_path}")
print(f"rows={len(rows)}")
